import re
import os
import json
import shutil
import requests
from time import sleep
from alpha import __log
import alpha._redis as REDIS
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from requests_toolbelt import MultipartEncoder
from alpha.settings import ALPHA_API_KEY, VITE_API_SERVER_URL, API_URL_PREFIX

def remove_special_characters(value):
    if isinstance(value, str):
        # Keep English letters, digits, spaces, and Korean characters
        return re.sub(r'[^A-Za-z0-9\s\uAC00-\uD7A3]', '', value)
    return value

def excel_download(request_dict):
    __log.debug(request_dict)
    URL = VITE_API_SERVER_URL + '/' + API_URL_PREFIX + request_dict['query']['url']
    
    multipart_params = MultipartEncoder(
        {
            "action": request_dict['query']['action'],
            "paging": json.dumps({'fullLoad': True}),
            "query": json.dumps(request_dict['query']['query']),
            "user": json.dumps(request_dict['user'])
        }
    )
    
    headers = {'Content-Type': multipart_params.content_type, 'Authorization' : f'Api-Key {ALPHA_API_KEY}'}

    response = requests.post(URL, data=multipart_params,  headers=headers)
    response_dict = response.json()

    template_path = '/www/alpha/apps/alpha_base/__excel_download_template/'
    result_path = '/tmp/'
    template = template_path + request_dict['query']['template'] + '.xlsx'

    new_filename = request_dict['query']['jobUUID'] + '.xlsx'
    new_filepath = result_path + new_filename

    shutil.copy(template, new_filepath)

    write_wb = load_workbook(new_filepath, data_only=True)
    rule_ws = write_wb.get_sheet_by_name('rule')

    rows = rule_ws.rows
    rules = {'start_row_num': 0, 'field_mapping': {}, 'field_format': {}, 'field_function': {}, 'field_alignment': {}}

    for row in rows:
        if row[0].value == 'start_row_num':
            rules['start_row_num'] = row[1].value
        if row[0].value == 'field_mapping':
            rules['field_mapping'][row[1].value] = row[2].value
        if row[0].value == 'field_format':
            rules['field_format'][row[1].value] = row[2].value
        if row[0].value == 'field_function':
            rules['field_function'][row[1].value] = row[2].value
        if row[0].value == 'field_alignment':
            rules['field_alignment'][row[1].value] = row[2].value

    write_ws = write_wb.get_sheet_by_name('data')

    for data in response_dict['data']:
        for key in data.keys():
            if key in rules['field_mapping']:
                try:
                    write_ws.cell(row=rules['start_row_num'], column=rules['field_mapping'][key]).value = data[key]
                except:
                    write_ws.cell(row=rules['start_row_num'], column=rules['field_mapping'][key]).value = 'due to an error in a text. this content has truncated.'
            if key in rules['field_format']:
                write_ws.cell(row=rules['start_row_num'], column=rules['field_mapping'][key]).number_format = rules['field_format'][key]
            if key in rules['field_function']:
                write_ws.cell(row=rules['start_row_num'], column=rules['field_mapping'][key]).value = eval(rules['field_function'][key])(data[key])
            if key in rules['field_alignment']:
                write_ws.cell(row=rules['start_row_num'], column=rules['field_mapping'][key]).alignment = Alignment(rules['field_alignment'][key])

        rules['start_row_num'] = rules['start_row_num'] + 1

    write_wb.active = write_wb['data']
    write_wb.remove_sheet(write_wb['rule'])
    write_wb.remove_sheet(write_wb['field_ref'])
    
    write_wb.save(new_filepath)
    write_wb.close()
    
    REDIS.saveResultFileToRedis(new_filepath)
    os.remove(new_filepath)

    return True, None, None, None


def wait(request_dict):
    # print('wait for' + str(request_dict['query']['seconds']))
    sleep(request_dict['query']['seconds'])
    return True, None, None, None